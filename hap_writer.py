"""Write the finished MasterTable to a clean, properly-typed .xlsx."""
import math, datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment

TEXT_COLS = {
 "Property ID","Property ID2","Contract Number","Last REAC 1","Last REAC 2","Last REAC 3",
 "Last REAC 1 Numeric","Last REAC 2 Numeric","Last REAC 3 Numeric","zip_code","zip4_code",
 "county_code","congressional_district_code","msa_code","owner_zip_code","owner_zip4_code",
 "mgmt_agent_zip_code","mgmt_agent_zip4_code","associated_financing_Number",
 "owner_main_phone_number_text","owner_main_fax_number_text","mgmt_agent_main_phone_number",
 "mgmt_agent_main_fax_number","property_phone_number","owner_participant_id","mgmt_agent_participant_id",
}
DATE_COLS = {"HAP Effective Date","HAP Expiration Date","Last REAC Date 1","Last REAC Date 2",
             "Last REAC Date 3","ownership_effective_date"}
NUM_COLS  = {"Total Units","Contract Units","Rent-to-FMR Ratio","property_total_unit_count",
             "1BR Rent","2BR Rent","3BR Rent","4BR Rent","0BR Rent","5BR+ Rent",
             "1BR UA","2BR UA","3BR UA","4BR UA","0BR UA","5BR+ UA",
             "0BR Units","1BR Units","2BR Units","3BR Units","4BR Units","5BR+ Units"}
EXCEL_EPOCH = dt.datetime(1899,12,30)

def _isblank(x):
    return x is None or (isinstance(x,float) and math.isnan(x)) or (isinstance(x,str) and x.strip()=="")

def to_date(x):
    if _isblank(x): return None
    if isinstance(x,(dt.datetime,dt.date)): return x
    if isinstance(x,pd.Timestamp): return x.to_pydatetime()
    s=str(x).strip()
    # excel serial number (possibly with fractional time)
    try:
        f=float(s)
        if 1 < f < 200000:
            return EXCEL_EPOCH + dt.timedelta(days=f)
    except ValueError:
        pass
    for fmt in ("%m/%d/%Y","%Y-%m-%d","%m/%d/%y","%Y-%m-%d %H:%M:%S"):
        try: return dt.datetime.strptime(s,fmt)
        except ValueError: continue
    return None

def to_num(x):
    if _isblank(x): return None
    try:
        f=float(str(x).replace(",","")); 
        return None if math.isnan(f) else f
    except ValueError: return None

def to_text(x):
    if _isblank(x): return None
    return str(x).strip()

def _write_sheet(wb, sheet_name, df, num_cols, date_cols, text_cols):
    ws=wb.create_sheet(sheet_name)
    cols=list(df.columns)
    hdr_font=Font(bold=True,color="FFFFFF")
    hdr_fill=PatternFill("solid",fgColor="1F4E78")
    header=[]
    for c in cols:
        cell=WriteOnlyCell(ws,value=c)
        cell.font=hdr_font; cell.fill=hdr_fill; cell.alignment=Alignment(horizontal="center")
        header.append(cell)
    ws.append(header)
    recs=df.to_dict("records")
    for rec in recs:
        row=[]
        for c in cols:
            v=rec[c]
            if c in date_cols:
                d=to_date(v); cell=WriteOnlyCell(ws,value=d)
                if d is not None: cell.number_format="m/d/yyyy"
            elif c in num_cols:
                nnum=to_num(v); cell=WriteOnlyCell(ws,value=nnum)
                if nnum is not None:
                    cell.number_format=("0.00" if c=="Rent-to-FMR Ratio" else "#,##0")
            elif c in text_cols:
                cell=WriteOnlyCell(ws,value=to_text(v)); cell.number_format="@"
            else:
                cell=WriteOnlyCell(ws,value=to_text(v))
            row.append(cell)
        ws.append(row)
    ws.freeze_panes="A2"
    return len(recs)

def write_master(df, path, sheet_name="MasterTable", extra_sheets=None,
                 num_cols=None, date_cols=None, text_cols=None):
    """extra_sheets: optional list of (name, df, num_cols, date_cols, text_cols)
    appended after the master. Column sets default to the master's sets, so
    default behaviour is unchanged."""
    wb=Workbook(write_only=True)
    n=_write_sheet(wb, sheet_name, df,
                   NUM_COLS if num_cols is None else num_cols,
                   DATE_COLS if date_cols is None else date_cols,
                   TEXT_COLS if text_cols is None else text_cols)
    for (nm, xdf, xnum, xdate, xtext) in (extra_sheets or []):
        _write_sheet(wb, nm, xdf, xnum, xdate, xtext)
    wb.save(path)
    return n
